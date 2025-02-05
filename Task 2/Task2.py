"""
CS3080 Python - Project 1 - Task 2.

This program will scrape three websites,
https://openaccess.thecvf.com/CVPR2022?day=all,
https://openaccess.thecvf.com/CVPR2023?day=all,
and https://openaccess.thecvf.com/CVPR2024?day=all,
and obtain the top three contributors (authors or researchers) in a
conference for the last three years, (2022, 2023, and 2024).
The results will be saved into an Excel spreadsheet
that shows how many times each author contributed
to a paper separated by the year.
"""
import openpyxl
from openpyxl.styles import Alignment, Font
import bs4
import requests

LINK_2022 = r'https://openaccess.thecvf.com/CVPR2022?day=all'
LINK_2023 = r'https://openaccess.thecvf.com/CVPR2023?day=all'
LINK_2024 = r'https://openaccess.thecvf.com/CVPR2024?day=all'


def find_contributors(link):
    """Totals Contributions for each year."""
    try:
        response = requests.get(link)
        response.raise_for_status()

    except Exception as e:
        print(f"Error {e}\nCheck Internet Connection")
        quit()

    soup = bs4.BeautifulSoup(response.text, 'html.parser')

    div = soup.find('div', {'id': 'content'})

    authors = list(div.select('input[type=hidden]'))

    author_counts = {}

    for element in authors:
        value = element.get("value")

        if value in author_counts:
            author_counts[value] += 1
        else:
            author_counts[value] = 1

    return author_counts


def count_totals(dict1, dict2, dict3):
    """Return the total for each contributor over the three years."""
    year_totals = {}

    def add_totals(key, values):

        if key in year_totals:
            year_totals[key][0].append(values)
            year_totals[key][1] += values
        else:
            year_totals[key] = [[values], values]

    for dict_in in [dict1, dict2, dict3]:
        for key, value in dict_in.items():
            add_totals(key, value)

    sorted_totals = sorted_totals = sorted([(key, values[0], values[1])
                                        for key, values in year_totals.items()],
                                        key=lambda x: x[2], reverse=True)

    return sorted_totals


def create_sheet(a_list):
    """Create the sheet with Top 3 Contributors."""
    top_con = openpyxl.Workbook()
    sheet = top_con.active
    sheet.title = "CVF Top 3 Contributors"

    # Add years and total to sheet
    start_year = 2022
    for i in range(4):
        if i < 3:
            cell = sheet.cell(row=i+2, column=1, value=start_year + i)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell = sheet.cell(row=i+2, column=1, value="Total")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Add names
    for i in range(3):
        cell = sheet.cell(row=1, column=i+2, value=a_list[i][0])
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add year 2022 values to sheet
    index = 2
    for key in a_list[:3]:
        i = 0
        for value in (key[1]):
            cell = sheet.cell(row=i+2, column=index, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            i += 1

        cell = sheet.cell(row=5, column=index, value=key[2])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        index += 1

    top_con.save('Task2.xlsx')


def main():
    """Run Main Program."""
    contributors_2022 = find_contributors(LINK_2022)
    contributors_2023 = find_contributors(LINK_2023)
    contributors_2024 = find_contributors(LINK_2024)

    totals = count_totals(contributors_2022,
                          contributors_2023,
                          contributors_2024)

    create_sheet(totals)


if __name__ == "__main__":
    main()

